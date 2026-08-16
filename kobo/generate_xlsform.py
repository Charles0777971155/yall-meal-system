"""
Generates the XLSForm that gets uploaded to Kobo.

Run this whenever config/indicators.py changes:

    python kobo/generate_xlsform.py

It writes kobo/YALL_MEAL_Indicator_Update.xlsx — upload that file directly
in Kobo via "New form" -> "Upload an XLSForm".

The form asks a coordinator to pick a project, then a community (filtered
to that project's real communities), then an indicator (filtered to that
project). What it asks next depends on the indicator's type:

  - "count" / "milestone" indicators -> a single numeric "value" field.
  - "percent" indicators -> "assessed" and "improved" fields.
  - "average" indicators -> a single "value" field, same as count, but
    logged repeatedly over time — the first entry becomes the baseline and
    every later one is compared back to it by the dashboard.

Always test the form in Kobo's preview after uploading, since XLSForm
engines can vary slightly in how they evaluate the conditional logic.
"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook
from config.indicators import PROJECTS, INDICATORS, COORDINATORS, COMMUNITIES

OUT_PATH = os.path.join(os.path.dirname(__file__), "YALL_MEAL_Indicator_Update.xlsx")

IS_PERCENT = "instance('indicator')/root/item[name=${indicator}]/kind='percent'"
IS_NOT_PERCENT = "instance('indicator')/root/item[name=${indicator}]/kind!='percent'"


def build():
    wb = Workbook()

    # ---- survey sheet: the questions, in order ----
    survey = wb.active
    survey.title = "survey"
    survey.append(["type", "name", "label", "required", "relevant", "choice_filter", "default", "constraint", "constraint_message"])
    survey.append(["select_one project", "project", "Project", "yes", "", "", "", "", ""])
    survey.append(["select_one community", "community", "Community", "yes", "", "project=${project}", "", "", ""])
    survey.append(["text", "group", "Group / cohort (optional, if you're running more than one group in this community)", "no", "", "", "", "", ""])
    survey.append(["select_one indicator", "indicator", "Indicator", "yes", "", "project=${project}", "", "", ""])
    survey.append(["decimal", "value", "New value recorded", "yes", IS_NOT_PERCENT, "", "", "", ""])
    survey.append(["integer", "assessed", "How many were assessed?", "yes", IS_PERCENT, "", "", "", ""])
    survey.append(["integer", "improved", "How many showed the improvement (or met the target behavior)?", "yes", IS_PERCENT, "", "", ".<=${assessed}", "Cannot be more than the number assessed"])
    survey.append(["date", "obs_date", "Date of observation", "yes", "", "", "today()", "", ""])
    survey.append(["geopoint", "location", "GPS location (optional — tap to capture on site)", "no", "", "", "", "", ""])
    survey.append(["image", "photo", "Photo evidence (optional)", "no", "", "", "", "", ""])
    survey.append(["text", "note", "Notes / context (optional)", "no", "", "", "", "", ""])
    survey.append(["select_one coordinator", "coordinator", "Recorded by", "yes", "", "", "", "", ""])

    # ---- choices sheet: the option lists referenced above ----
    choices = wb.create_sheet("choices")
    choices.append(["list_name", "name", "label", "project", "kind"])

    for p in PROJECTS:
        choices.append(["project", p["id"], p["name"], "", ""])

    for project_id, community_list in COMMUNITIES.items():
        for community in community_list:
            choices.append(["community", community.lower().replace(" ", "_").replace("-", "_"), community, project_id, ""])

    for ind in INDICATORS:
        # "average" indicators behave like "count" in the form (a single value field);
        # the "kind" column here only distinguishes percent vs everything else.
        kind_for_relevant = "percent" if ind["type"] == "percent" else "count"
        label = f"{ind['name']} ({ind['unit']})" if ind["type"] != "percent" else f"{ind['name']} (%)"
        choices.append(["indicator", ind["id"], label, ind["project_id"], kind_for_relevant])

    for c in COORDINATORS:
        choices.append(["coordinator", c["username"], c["name"], "", ""])

    # ---- settings sheet: form metadata ----
    settings = wb.create_sheet("settings")
    settings.append(["form_title", "form_id", "allow_choice_duplicates"])
    settings.append(["YALL M&E Indicator Update", "yall_meal_indicator_update", "true"])

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    n_communities = sum(len(v) for v in COMMUNITIES.values())
    print(f"  {len(PROJECTS)} projects, {len(INDICATORS)} indicators, {n_communities} community entries, {len(COORDINATORS)} coordinators")


if __name__ == "__main__":
    build()
